from __future__ import annotations

from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from administration.models import Supplier
from administration.services.supplier_imports import (
    SUPPLIER_IMPORT_TEMPLATE_HEADERS,
    SupplierImportError,
    import_suppliers_from_workbook,
)


class SupplierImportServiceTests(TestCase):
    def _build_workbook(self, rows: list[list[str]]) -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def test_creates_and_updates_suppliers(self) -> None:
        Supplier.objects.create(name="Proveedor antiguo", tax_id="100")
        buffer = self._build_workbook(
            [
                ["Nombre", "CC/NIT", "Correo", "Teléfono"],
                ["Proveedor Uno", "123", "uno@example.com", "3000000"],
                ["Proveedor Dos", "100", "dos@example.com", ""],
            ]
        )

        result = import_suppliers_from_workbook(buffer)

        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.updated_count, 1)
        self.assertEqual(Supplier.objects.count(), 2)
        supplier_new = Supplier.objects.get(tax_id="123")
        supplier_updated = Supplier.objects.get(tax_id="100")
        self.assertEqual(supplier_new.contact_email, "uno@example.com")
        self.assertEqual(supplier_updated.contact_email, "dos@example.com")

    def test_requires_mandatory_columns(self) -> None:
        buffer = self._build_workbook(
            [
                ["Nombre"],
                ["Proveedor"],
            ]
        )

        with self.assertRaises(SupplierImportError):
            import_suppliers_from_workbook(buffer)

    def test_collects_row_errors_when_missing_required_fields(self) -> None:
        buffer = self._build_workbook(
            [
                ["Nombre", "CC/NIT"],
                ["Proveedor Incompleto", ""],
                ["", "9999"],
                ["Completo", "5555"],
            ]
        )

        result = import_suppliers_from_workbook(buffer)

        self.assertEqual(result.errors[0].row_number, 2)
        self.assertEqual(result.errors[1].row_number, 3)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(Supplier.objects.filter(tax_id="5555").count(), 1)


class SupplierImportTemplateViewTests(TestCase):
    def setUp(self) -> None:
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            cedula="1001",
            password="testpass123",
            nombres="Staff",
            apellidos="Example",
            telefono="3000000",
            is_staff=True,
        )
        self.client.force_login(self.user)

    def test_template_download_contains_expected_headers(self) -> None:
        response = self.client.get(reverse('administration:purchases_supplier_import_template'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        self.assertEqual(list(header_row), SUPPLIER_IMPORT_TEMPLATE_HEADERS)
