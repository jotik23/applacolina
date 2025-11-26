from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from openpyxl import Workbook

from administration.models import Sale, SaleProductType, Supplier
from administration.services.sale_imports import import_sales_from_workbook
from personal.models import UserProfile


class SaleImportServiceTestCase(TestCase):
    def setUp(self) -> None:
        self.actor = UserProfile.objects.create_user(
            cedula="11111111",
            password="secret",
            nombres="Coordinador",
            apellidos="General",
            telefono="3000000000",
            is_staff=True,
        )
        self.seller = UserProfile.objects.create_user(
            cedula="22222222",
            password="secret",
            nombres="Osnaider",
            apellidos="Ventas",
            telefono="3010000000",
            is_staff=True,
        )
        self.billing_user = UserProfile.objects.create_user(
            cedula="33333333",
            password="secret",
            nombres="Kerlys",
            apellidos="Contable",
            telefono="3020000000",
            is_staff=True,
        )

    def _build_workbook(self) -> BytesIO:
        workbook = Workbook()
        sales_sheet = workbook.active
        sales_sheet.title = "Ventas"
        sales_sheet.append(
            [
                "Fecha",
                "Num. Factura",
                "Enviado a la DIAN",
                "Bodega",
                "Cliente",
                "Valor Total",
                "Saldo Actual",
                "Cantidad Jumbo",
                "Precio Jumbo",
                "Cantidad AAA",
                "Precio AAA",
                "Comentarios",
                "Vendedor",
                "Facturador",
            ]
        )
        sales_sheet.append(
            [
                date(2025, 1, 15),
                "FV - 1001",
                "Enviada",
                "Montería",
                "Cliente Uno",
                50000,
                0,
                2,
                25000,
                0,
                0,
                "Entrega prioritaria",
                "Osnaider",
                "Kerlys",
            ]
        )
        sales_sheet.append(
            [
                date(2025, 1, 16),
                "FV-1002",
                "No",
                "Tierralta",
                "Cliente Nuevo",
                30000,
                30000,
                0,
                0,
                30,
                1000,
                "",
                "Desconocido",
                "",
            ]
        )
        payments_sheet = workbook.create_sheet("Abonos")
        payments_sheet.append(["Num. Factura", "Fecha Abono", "Monto Abonado", "Modalidad"])
        payments_sheet.append(["FV - 1001", date(2025, 1, 20), 50000, "Transferencia"])
        third_sheet = workbook.create_sheet("Terceros")
        third_sheet.append(["Código", "Nombre", "Identificación"])
        third_sheet.append([1, "Cliente Uno", 900123456])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def test_import_creates_sales_suppliers_and_payments(self):
        stream = self._build_workbook()
        result = import_sales_from_workbook(stream, actor=self.actor)

        self.assertEqual(result.created_sales, 2)
        self.assertEqual(result.updated_sales, 0)
        self.assertEqual(result.registered_payments, 1)
        self.assertEqual(result.created_suppliers, 2)
        self.assertFalse(result.issues)

        sale_one = Sale.objects.get(invoice_number="FV-1001")
        self.assertTrue(sale_one.sent_to_dian)
        self.assertEqual(sale_one.seller, self.seller)
        self.assertEqual(sale_one.confirmed_by, self.billing_user)
        self.assertEqual(sale_one.customer.name, "Cliente Uno")
        self.assertEqual(sale_one.payments_total, Decimal("50000"))
        self.assertEqual(sale_one.items.count(), 1)
        item = sale_one.items.first()
        assert item is not None
        self.assertEqual(item.product_type, SaleProductType.JUMBO)
        self.assertEqual(item.quantity, Decimal("2"))

        sale_two = Sale.objects.get(invoice_number="FV-1002")
        self.assertFalse(sale_two.sent_to_dian)
        self.assertEqual(sale_two.seller, self.actor)
        self.assertEqual(sale_two.payment_condition, Sale.PaymentCondition.CREDIT)
        self.assertEqual(sale_two.items.count(), 1)
        second_item = sale_two.items.first()
        assert second_item is not None
        self.assertEqual(second_item.product_type, SaleProductType.TRIPLE_A)
        self.assertEqual(second_item.quantity, Decimal("30"))
        self.assertTrue(sale_two.customer.tax_id.startswith("AUTO-"))
        self.assertEqual(Supplier.objects.count(), 2)
