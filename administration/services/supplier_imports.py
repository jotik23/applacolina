from __future__ import annotations

from dataclasses import dataclass
from typing import Any
import re
import unicodedata

from django.db import transaction
from openpyxl import load_workbook

from administration.models import Supplier


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^\w\s]", "_", text, flags=re.ASCII)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def _alias_set(*values: str) -> set[str]:
    return {_normalize_header(value) for value in values if value}


class SupplierImportError(Exception):
    """Raised when the uploaded file is not a valid supplier import."""


@dataclass
class SupplierImportRowError:
    row_number: int
    message: str


@dataclass
class SupplierImportResult:
    created_count: int
    updated_count: int
    processed_rows: int
    errors: list[SupplierImportRowError]


REQUIRED_FIELDS = {"name", "tax_id"}

SUPPLIER_IMPORT_TEMPLATE_HEADERS = [
    "Nombre",
    "CC/NIT",
    "Contacto",
    "Correo",
    "Teléfono",
    "Dirección",
    "Ciudad",
    "Identificación titular",
    "Nombre titular",
    "Tipo de cuenta",
    "Número de cuenta",
    "Banco",
]

SUPPLIER_IMPORT_TEMPLATE_SAMPLE_ROW = [
    "Proveedor Demo",
    "900123456",
    "María Pérez",
    "proveedor@example.com",
    "3001234567",
    "Km 5 vía principal",
    "Bogotá",
    "900123456",
    "María Pérez",
    "Ahorros",
    "0011223344",
    "Banco Ejemplo",
]

FIELD_ALIASES: dict[str, set[str]] = {
    "name": _alias_set("nombre", "name", "razon social", "razón social", "razon_social", "razon"),
    "tax_id": _alias_set(
        "ccnit",
        "cc/nit",
        "cc_nit",
        "nit",
        "documento",
        "documento identificacion",
        "identificacion",
        "identificación",
        "tax_id",
        "numero documento",
        "numero_documento",
    ),
    "contact_name": _alias_set("contacto", "contact_name", "nombre_contacto"),
    "contact_email": _alias_set("correo", "email", "correo_electronico"),
    "contact_phone": _alias_set("telefono", "phone", "celular", "telefono_contacto"),
    "address": _alias_set("direccion", "address", "direccion correspondencia"),
    "city": _alias_set("ciudad", "city"),
    "account_holder_id": _alias_set("identificacion titular", "documento titular", "titular_id"),
    "account_holder_name": _alias_set("nombre titular", "titular_nombre"),
    "account_type": _alias_set("tipo cuenta", "tipo de cuenta", "account_type"),
    "account_number": _alias_set("numero cuenta", "account_number"),
    "bank_name": _alias_set("banco", "bank", "bank_name"),
}


def import_suppliers_from_workbook(file_obj) -> SupplierImportResult:
    """Parse the uploaded workbook and persist Supplier instances."""
    file_obj.seek(0)
    try:
        workbook = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive
        raise SupplierImportError("No fue posible leer el archivo. Verifica que sea un .xlsx válido.") from exc

    sheet = workbook.active
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not header_cells:
        raise SupplierImportError("El archivo está vacío.")

    header_map = _build_header_map(header_cells)
    missing = REQUIRED_FIELDS - set(header_map)
    if missing:
        titles = ", ".join(sorted(_field_verbose_name(field) for field in missing))
        raise SupplierImportError(f"El archivo debe incluir las columnas obligatorias: {titles}.")

    created = updated = processed = 0
    errors: list[SupplierImportRowError] = []
    account_type_lookup = _build_account_type_lookup()

    with transaction.atomic():
        for row in sheet.iter_rows(min_row=2):
            row_number = row[0].row
            row_values = _extract_row_values(row, header_map)
            if not any(row_values.values()):
                continue
            processed += 1
            name = row_values.get("name", "").strip()
            tax_id = row_values.get("tax_id", "").strip()
            if not name or not tax_id:
                errors.append(
                    SupplierImportRowError(
                        row_number=row_number,
                        message="Las columnas Nombre y Identificación son obligatorias.",
                    )
                )
                continue
            payload = {
                "name": name,
                "tax_id": tax_id,
                "contact_name": row_values.get("contact_name", ""),
                "contact_email": row_values.get("contact_email", ""),
                "contact_phone": row_values.get("contact_phone", ""),
                "address": row_values.get("address", ""),
                "city": row_values.get("city", ""),
                "account_holder_id": row_values.get("account_holder_id", ""),
                "account_holder_name": row_values.get("account_holder_name", ""),
                "account_number": row_values.get("account_number", ""),
                "bank_name": row_values.get("bank_name", ""),
            }
            account_type_raw = row_values.get("account_type", "")
            if account_type_raw:
                normalized = _normalize_header(account_type_raw)
                payload["account_type"] = account_type_lookup.get(normalized, "")
            supplier, created_flag = Supplier.objects.update_or_create(
                tax_id=tax_id,
                defaults=payload,
            )
            if created_flag:
                created += 1
            else:
                updated += 1

    return SupplierImportResult(
        created_count=created,
        updated_count=updated,
        processed_rows=processed,
        errors=errors,
    )


def _build_header_map(cells) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for idx, cell in enumerate(cells, start=1):
        value = cell.value
        if not isinstance(value, str):
            continue
        normalized = _normalize_header(value)
        for field, aliases in FIELD_ALIASES.items():
            if normalized in aliases and field not in header_map:
                header_map[field] = idx
    return header_map


def _extract_row_values(row, header_map: dict[str, int]) -> dict[str, str]:
    values: dict[str, str] = {}
    for field, column_index in header_map.items():
        cell = row[column_index - 1]
        values[field] = _stringify_value(cell.value)
    return values


def _stringify_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def _build_account_type_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for value, _ in Supplier.ACCOUNT_TYPE_CHOICES:
        normalized = _normalize_header(value)
        lookup[normalized] = value
    lookup["ahorro"] = "ahorros"
    lookup["corriente"] = "corriente"
    return lookup


def _field_verbose_name(field: str) -> str:
    if field == "name":
        return "Nombre"
    if field == "tax_id":
        return "Número de identificación"
    return field
