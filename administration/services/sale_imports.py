from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import re
from typing import Any, Dict, Iterable, Tuple
import unicodedata

from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import load_workbook

from administration.models import Sale, SaleItem, SalePayment, SaleProductType, Supplier
from administration.services.sales import refresh_sale_payment_state
from production.models import EggDispatchDestination


class SaleImportError(Exception):
    """Raised when the workbook does not have the required structure."""


@dataclass
class SaleImportIssue:
    sheet: str
    row_number: int
    message: str
    reference: str | None = None


@dataclass
class SaleImportResult:
    created_sales: int = 0
    updated_sales: int = 0
    created_suppliers: int = 0
    registered_payments: int = 0
    issues: list[SaleImportIssue] | None = None

    def __post_init__(self) -> None:
        if self.issues is None:
            self.issues = []


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^\w\s]", "_", text, flags=re.ASCII)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


SALE_FIELD_ALIASES: dict[str, set[str]] = {
    "date": {"fecha"},
    "invoice_number": {"num_factura", "numero_factura"},
    "sent_to_dian": {"enviado_a_la_dian", "enviada_a_la_dian"},
    "warehouse_destination": {"bodega", "destino"},
    "customer_name": {"cliente"},
    "total_amount": {"valor_total", "total_factura"},
    "balance": {"saldo_actual", "saldo"},
    "notes": {"comentarios", "notas"},
    "seller_name": {"vendedor"},
    "invoiced_by": {"facturador", "facturado_por"},
}

PRODUCT_COLUMN_ALIASES: dict[str, dict[str, set[str]]] = {
    SaleProductType.JUMBO: {"quantity": {"cantidad_jumbo"}, "price": {"precio_jumbo"}},
    SaleProductType.TRIPLE_A: {"quantity": {"cantidad_aaa"}, "price": {"precio_aaa"}},
    SaleProductType.DOUBLE_A: {"quantity": {"cantidad_aa"}, "price": {"precio_aa"}},
    SaleProductType.SINGLE_A: {"quantity": {"cantidad_a"}, "price": {"precio_a"}},
    SaleProductType.B: {"quantity": {"cantidad_b"}, "price": {"precio_b"}},
    SaleProductType.C: {"quantity": {"cantidad_c"}, "price": {"precio_c"}},
    SaleProductType.D: {"quantity": {"cantidad_d"}, "price": {"precio_d"}},
    SaleProductType.HEN: {"quantity": {"cantidad_gallinas"}, "price": {"precio_gallinas"}},
    SaleProductType.HEN_MANURE: {"quantity": {"cantidad_gallinaza"}, "price": {"precio_gallinaza"}},
}

PAYMENT_FIELD_ALIASES: dict[str, set[str]] = {
    "invoice_number": {"num_factura", "numero_factura"},
    "date": {"fecha_abono", "fecha"},
    "amount": {"monto_abonado", "valor_abono", "valor"},
    "method": {"modalidad", "metodo"},
}

DESTINATION_LOOKUP: dict[str, str] = {}
for code, label in EggDispatchDestination.choices:
    for key in {
        code,
        code.lower(),
        _normalize_header(code),
        label,
        label.lower(),
        _normalize_header(label),
    }:
        if key:
            DESTINATION_LOOKUP.setdefault(key, code)

UserModel = get_user_model()


def import_sales_from_workbook(file_obj, *, actor=None) -> SaleImportResult:
    """Main entry point used by the UI to import sales, suppliers and payments."""
    file_obj.seek(0)
    try:
        workbook = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive guard
        raise SaleImportError("No se pudo leer el archivo. Verifica que sea un .xlsx válido.") from exc

    sheet_lookup = {_normalize_header(name): name for name in workbook.sheetnames}
    sales_sheet = _get_sheet(workbook, sheet_lookup, "ventas")
    payments_sheet = _get_sheet(workbook, sheet_lookup, "abonos")
    third_sheet = _get_sheet(workbook, sheet_lookup, "terceros")

    existing_tax_ids = set(Supplier.objects.values_list("tax_id", flat=True))
    supplier_lookup = {_normalize_name_text(s.name): s for s in Supplier.objects.all()}

    result = SaleImportResult()

    supplier_created, supplier_issues = _import_third_parties(third_sheet, supplier_lookup, existing_tax_ids)
    result.created_suppliers += supplier_created
    result.issues.extend(supplier_issues)

    sales_created, sales_updated, additional_suppliers, sale_lookup, sale_issues = _import_sales_sheet(
        sales_sheet,
        supplier_lookup,
        existing_tax_ids,
        actor=actor,
    )
    result.created_sales += sales_created
    result.updated_sales += sales_updated
    result.created_suppliers += additional_suppliers
    result.issues.extend(sale_issues)

    payments_registered, payment_issues = _import_payment_sheet(payments_sheet, sale_lookup)
    result.registered_payments += payments_registered
    result.issues.extend(payment_issues)

    return result


def _import_third_parties(sheet, supplier_lookup: dict[str, Supplier], existing_tax_ids: set[str]) -> tuple[int, list[SaleImportIssue]]:
    header = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not header:
        raise SaleImportError("La hoja Terceros no tiene encabezados.")
    created = 0
    issues: list[SaleImportIssue] = []
    for row in sheet.iter_rows(min_row=2):
        name_raw = row[1].value if len(row) > 1 else None
        tax_raw = row[2].value if len(row) > 2 else None
        code_value = row[0].value if row else None
        name = _stringify(name_raw)
        if not name:
            continue
        tax_id = _normalize_tax_id(tax_raw)
        if not tax_id and code_value is not None:
            tax_id = f"AUTO-{_normalize_tax_id(code_value) or _normalize_invoice_number(code_value)}"
        if not tax_id:
            issues.append(
                SaleImportIssue(
                    sheet="Terceros",
                    row_number=row[0].row,
                    message="No se pudo crear el tercero por falta de identificación.",
                    reference=name,
                )
            )
            continue
        supplier, created_flag = Supplier.objects.update_or_create(
            tax_id=tax_id,
            defaults={"name": name},
        )
        if created_flag:
            created += 1
            existing_tax_ids.add(tax_id)
        supplier_lookup[_normalize_name_text(name)] = supplier
    return created, issues


def _import_sales_sheet(
    sheet,
    supplier_lookup: dict[str, Supplier],
    existing_tax_ids: set[str],
    *,
    actor=None,
) -> tuple[int, int, int, dict[str, Sale], list[SaleImportIssue]]:
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not header_cells:
        raise SaleImportError("La hoja Ventas está vacía.")

    header_index = _build_header_index(header_cells)
    field_columns = _resolve_columns(header_index, SALE_FIELD_ALIASES)
    product_columns = _resolve_product_columns(header_index)

    created = updated = created_suppliers = 0
    sale_lookup: dict[str, Sale] = {}
    issues: list[SaleImportIssue] = []

    for row in sheet.iter_rows(min_row=2):
        row_number = row[0].row
        sale_date = _coerce_date(_cell_value(row, field_columns.get("date")))
        raw_invoice = _cell_value(row, field_columns.get("invoice_number"))
        invoice = _normalize_invoice_number(raw_invoice)
        customer_name = _stringify(_cell_value(row, field_columns.get("customer_name")))
        if not invoice or not sale_date or not customer_name:
            if not invoice and not customer_name and not sale_date:
                continue
            issues.append(
                SaleImportIssue(
                    sheet="Ventas",
                    row_number=row_number,
                    message="Faltan datos obligatorios (fecha, factura o cliente).",
                )
            )
            continue
        items = _build_items_payload(row, product_columns)
        if not items:
            issues.append(
                SaleImportIssue(
                    sheet="Ventas",
                    row_number=row_number,
                    message="La fila no incluye cantidades válidas para productos.",
                    reference=invoice,
                )
            )
            continue

        customer, created_flag = _resolve_customer(customer_name, supplier_lookup, existing_tax_ids)
        if created_flag:
            created_suppliers += 1

        seller_name = _stringify(_cell_value(row, field_columns.get("seller_name")))
        seller = _match_user(seller_name) or actor
        if not seller:
            issues.append(
                SaleImportIssue(
                    sheet="Ventas",
                    row_number=row_number,
                    message="No se encontró un vendedor para asignar la venta.",
                    reference=invoice,
                )
            )
            continue

        invoiced_by_name = _stringify(_cell_value(row, field_columns.get("invoiced_by")))
        invoiced_by = _match_user(invoiced_by_name) or seller
        total_value = _to_decimal(_cell_value(row, field_columns.get("total_amount")))
        balance = _to_decimal(_cell_value(row, field_columns.get("balance")))
        notes = _stringify(_cell_value(row, field_columns.get("notes")))
        sent_to_dian = _normalize_bool(_cell_value(row, field_columns.get("sent_to_dian")))
        destination_value = _stringify(_cell_value(row, field_columns.get("warehouse_destination")))
        destination = _resolve_destination(destination_value)

        subtotal = sum((payload["subtotal"] for payload in items.values()), Decimal("0.00"))
        declared_total = total_value if total_value > Decimal("0") else subtotal
        discount = max(Decimal("0.00"), subtotal - declared_total).quantize(Decimal("0.01"))
        status = Sale.Status.PAID if balance <= Decimal("0.00") else Sale.Status.CONFIRMED
        payment_condition = Sale.PaymentCondition.CREDIT if balance > Decimal("0.00") else Sale.PaymentCondition.CASH
        payment_due_date = sale_date if payment_condition == Sale.PaymentCondition.CREDIT else None

        existing_sale = (
            Sale.objects.select_related("customer")
            .filter(invoice_number=invoice)
            .first()
        )
        is_new = existing_sale is None
        sale = existing_sale or Sale(invoice_number=invoice)

        with transaction.atomic():
            sale.date = sale_date
            sale.customer = customer
            sale.seller = seller
            sale.status = status
            sale.payment_condition = payment_condition
            sale.payment_due_date = payment_due_date
            sale.notes = notes
            sale.sent_to_dian = sent_to_dian
            sale.warehouse_destination = destination or ""
            sale.discount_amount = discount
            sale.invoice_number = invoice
            if status in (Sale.Status.CONFIRMED, Sale.Status.PAID):
                sale.confirmed_at = sale.confirmed_at or timezone.now()
                sale.confirmed_by = invoiced_by
            else:
                sale.confirmed_at = None
                sale.confirmed_by = None
            sale.save()
            SaleItem.objects.filter(sale=sale).delete()
            sale_items = [
                SaleItem(
                    sale=sale,
                    product_type=product_type,
                    quantity=payload["quantity"],
                    unit_price=payload["unit_price"],
                    subtotal=payload["subtotal"],
                )
                for product_type, payload in items.items()
            ]
            SaleItem.objects.bulk_create(sale_items)
            refresh_sale_payment_state(sale)

        sale_lookup[invoice] = sale
        if is_new:
            created += 1
        else:
            updated += 1
    return created, updated, created_suppliers, sale_lookup, issues


def _import_payment_sheet(sheet, sale_lookup: dict[str, Sale]) -> tuple[int, list[SaleImportIssue]]:
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not header_cells:
        raise SaleImportError("La hoja Abonos está vacía.")
    header_index = _build_header_index(header_cells)
    field_columns = _resolve_columns(header_index, PAYMENT_FIELD_ALIASES)
    issues: list[SaleImportIssue] = []
    registered = 0

    for row in sheet.iter_rows(min_row=2):
        row_number = row[0].row
        invoice = _normalize_invoice_number(_cell_value(row, field_columns.get("invoice_number")))
        if not invoice:
            continue
        sale = sale_lookup.get(invoice) or Sale.objects.filter(invoice_number=invoice).first()
        if not sale:
            issues.append(
                SaleImportIssue(
                    sheet="Abonos",
                    row_number=row_number,
                    message="No existe una venta con el número de factura indicado.",
                    reference=invoice,
                )
            )
            continue
        payment_date = _coerce_date(_cell_value(row, field_columns.get("date"))) or sale.date
        amount = _to_decimal(_cell_value(row, field_columns.get("amount")))
        if amount <= Decimal("0.00"):
            continue
        method = _resolve_payment_method(_cell_value(row, field_columns.get("method")))
        with transaction.atomic():
            exists = SalePayment.objects.filter(
                sale=sale,
                date=payment_date,
                amount=amount,
                method=method,
            ).exists()
            if exists:
                continue
            SalePayment.objects.create(
                sale=sale,
                date=payment_date,
                amount=amount,
                method=method,
                notes="Importación de Excel",
            )
            registered += 1
            refresh_sale_payment_state(sale)
    return registered, issues


def _build_items_payload(row, product_columns: dict[str, dict[str, int]]) -> Dict[str, Dict[str, Decimal]]:
    items: Dict[str, Dict[str, Decimal]] = {}
    for product_type, columns in product_columns.items():
        quantity = _to_decimal(_cell_value(row, columns.get("quantity")))
        unit_price = _to_decimal(_cell_value(row, columns.get("price")))
        if quantity <= Decimal("0.00") or unit_price <= Decimal("0.00"):
            continue
        subtotal = (quantity * unit_price).quantize(Decimal("0.01"))
        items[product_type] = {
            "quantity": quantity,
            "unit_price": unit_price,
            "subtotal": subtotal,
        }
    return items


def _resolve_customer(
    name: str,
    supplier_lookup: dict[str, Supplier],
    existing_tax_ids: set[str],
) -> tuple[Supplier, bool]:
    normalized = _normalize_name_text(name)
    supplier = supplier_lookup.get(normalized)
    if supplier:
        return supplier, False
    supplier = Supplier.objects.filter(name__iexact=name).first()
    if supplier:
        supplier_lookup[normalized] = supplier
        return supplier, False
    tax_id = _generate_tax_id(name, existing_tax_ids)
    supplier = Supplier.objects.create(name=name, tax_id=tax_id)
    supplier_lookup[normalized] = supplier
    existing_tax_ids.add(tax_id)
    return supplier, True


def _generate_tax_id(name: str, existing_tax_ids: set[str]) -> str:
    base = re.sub(r"[^A-Z0-9]", "", _normalize_invoice_number(name)) or "TERCERO"
    base = base[:12]
    candidate = f"AUTO-{base}"
    counter = 1
    while candidate in existing_tax_ids:
        candidate = f"AUTO-{base}{counter}"
        counter += 1
    return candidate


def _match_user(value: str | None):
    if not value:
        return None
    tokens = [token for token in value.split() if token]
    if not tokens:
        return None
    query = Q(nombres__iexact=value) | Q(apellidos__iexact=value)
    for token in tokens:
        query |= Q(nombres__icontains=token) | Q(apellidos__icontains=token)
    if len(tokens) >= 2:
        query |= Q(nombres__icontains=tokens[0], apellidos__icontains=tokens[-1])
    return UserModel.objects.filter(query).order_by("apellidos", "nombres").first()


def _normalize_invoice_number(value: Any) -> str:
    text = _stringify(value)
    if not text:
        return ""
    text = text.replace(" ", "").replace("\u00a0", "")
    return text.upper()


def _normalize_name_text(value: str) -> str:
    normalized = _normalize_header(value)
    return normalized.replace("_", "")


def _normalize_tax_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)).strip()
    text = _stringify(value)
    if not text:
        return ""
    text = text.replace(".", "").replace(",", "")
    if text.isdigit():
        return text
    return text.upper()


def _normalize_bool(value: Any) -> bool:
    text = _stringify(value)
    if not text:
        return False
    normalized = unicodedata.normalize("NFKD", text).lower()
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.replace(" ", "")
    return any(keyword in normalized for keyword in ("si", "true", "enviado", "enviada", "1"))


def _resolve_destination(value: str | None) -> str | None:
    if not value:
        return None
    normalized = _normalize_header(value)
    if normalized in DESTINATION_LOOKUP:
        return DESTINATION_LOOKUP[normalized]
    lower_value = value.lower()
    return DESTINATION_LOOKUP.get(lower_value)


def _resolve_payment_method(value: Any) -> str:
    text = _stringify(value).lower()
    if "transfer" in text:
        return SalePayment.Method.TRANSFER
    return SalePayment.Method.CASH


def _coerce_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = parse_date(str(value))
    return parsed


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    text = str(value).strip()
    if not text:
        return Decimal("0.00")
    text = text.replace(",", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _get_sheet(workbook, sheet_lookup: dict[str, str], key: str):
    normalized = _normalize_header(key)
    sheet_name = sheet_lookup.get(normalized)
    if not sheet_name:
        raise SaleImportError(f"No se encontró la hoja requerida: {key}.")
    return workbook[sheet_name]


def _build_header_index(cells: Iterable) -> dict[str, int]:
    header_index: dict[str, int] = {}
    for idx, cell in enumerate(cells, start=1):
        normalized = _normalize_header(cell.value)
        if normalized and normalized not in header_index:
            header_index[normalized] = idx
    return header_index


def _resolve_columns(header_index: dict[str, int], aliases: dict[str, set[str]]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for field, options in aliases.items():
        for candidate in options:
            column = header_index.get(candidate)
            if column:
                columns[field] = column
                break
    return columns


def _resolve_product_columns(header_index: dict[str, int]) -> dict[str, dict[str, int]]:
    mapping: dict[str, dict[str, int]] = {}
    for product_type, sections in PRODUCT_COLUMN_ALIASES.items():
        quantity_column = _first_match(header_index, sections.get("quantity", set()))
        price_column = _first_match(header_index, sections.get("price", set()))
        if quantity_column and price_column:
            mapping[product_type] = {"quantity": quantity_column, "price": price_column}
    return mapping


def _first_match(header_index: dict[str, int], options: Iterable[str]) -> int | None:
    for candidate in options:
        column = header_index.get(candidate)
        if column:
            return column
    return None


def _cell_value(row, column_index: int | None):
    if not column_index:
        return None
    if column_index - 1 >= len(row):
        return None
    return row[column_index - 1].value
